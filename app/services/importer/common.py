from __future__ import annotations

from datetime import datetime, date
from typing import Any, Iterable, List, Dict

from openpyxl import load_workbook

from ...utils.normalize import normalize_text

# re-export for convenience
__all__ = [
	"read_sheet_rows",
	"row_to_dict",
	"parse_date",
	"parse_float",
	"parse_int",
	"normalize_header",
]


def normalize_header(h: str) -> str:
	return normalize_text(h)


def parse_float(value: Any) -> float | None:
	if value is None:
		return None
	if isinstance(value, (int, float)):
		return float(value)
	s = str(value).strip().replace(" ", "")
	s = s.replace(".", "").replace(",", ".") if s.count(",") == 1 and s.count(".") > 1 else s
	try:
		return float(s)
	except Exception:
		return None


def parse_int(value: Any) -> int | None:
	if value is None:
		return None
	try:
		return int(float(value))
	except Exception:
		return None


def parse_date(value: Any) -> date | None:
	if value is None:
		return None
	# normalize datetime to date
	if isinstance(value, datetime):
		return value.date()
	if isinstance(value, date):
		return value
	s = str(value).strip()
	for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
		try:
			return datetime.strptime(s, fmt).date()
		except Exception:
			continue
	return None


def read_sheet_rows(file_path: str) -> tuple[list[str], list[list[Any]]]:
	wb = load_workbook(filename=file_path, data_only=True)
	ws = wb.active
	headers: list[str] = []
	rows: list[list[Any]] = []
	for i, row in enumerate(ws.iter_rows(values_only=True)):
		cells = [c if c is not None else None for c in row]
		if i == 0:
			headers = [normalize_header(str(h)) if h is not None else "" for h in cells]
			# DEBUG: print normalized headers
			try:
				print("[HEADER DEBUG] raw:", [str(h) if h is not None else '' for h in cells])
				print("[HEADER DEBUG] normalized:", headers)
			except Exception:
				pass
			continue
		rows.append(cells)
	return headers, rows


def row_to_dict(headers: list[str], row: list[Any]) -> Dict[str, Any]:
	data: Dict[str, Any] = {}
	for idx, h in enumerate(headers):
		if not h:
			continue
		data[h] = row[idx] if idx < len(row) else None
	return data
