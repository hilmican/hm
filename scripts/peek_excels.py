#!/usr/bin/env python3
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook

from app.services.importer.common import normalize_header
from app.services.importer.bizim import read_bizim_file
from app.services.importer.kargo import read_kargo_file


ROOT = Path(__file__).resolve().parents[1]
FOLDERS = [
	(ROOT / "bizimexcellerimiz", "bizim"),
	(ROOT / "kargocununexcelleri", "kargo"),
]


def peek_headers(path: Path):
	wb = load_workbook(filename=str(path), data_only=True)
	ws = wb.active
	rows = list(ws.iter_rows(values_only=True))
	headers = [normalize_header(str(h)) if h is not None else '' for h in rows[0]] if rows else []
	return headers, rows[1:6]


def counts_for_kargo(records):
	keys = Counter()
	for r in records:
		for k in ("shipment_date","payment_amount","name","address","city","tracking_no"):
			if r.get(k):
				keys[k]+=1
		for k in ("fee_komisyon","fee_hizmet","fee_kargo","fee_iade","fee_erken_odeme","payment_method","delivery_date"):
			if r.get(k):
				keys[k]+=1
	return keys


def counts_for_bizim(records):
	keys = Counter()
	for r in records:
		for k in ("name","phone","address","city","item_name","quantity","unit_price","total_amount","shipment_date","tracking_no"):
			if r.get(k):
				keys[k]+=1
	return keys


def main():
	for folder, source in FOLDERS:
		if not folder.exists():
			print(f"Skip missing folder: {folder}")
			continue
		print(f"\n=== Folder: {folder.name} (source={source}) ===")
		for p in sorted(folder.glob("*.xlsx"))[:20]:
			headers, first_rows = peek_headers(p)
			print(f"\n-- {p.name}")
			print("Headers:", headers)
			try:
				records = read_bizim_file(str(p)) if source=="bizim" else read_kargo_file(str(p))
				print(f"Mapped records: {len(records)}")
				for i, rec in enumerate(records[:3], start=1):
					filled = {k:v for k,v in rec.items() if v not in (None, '', 0)}
					print(f"Sample {i} keys:", sorted(filled.keys()))
				cnt = counts_for_bizim(records) if source=="bizim" else counts_for_kargo(records)
				if cnt:
					print("Presence counts:", dict(cnt))
			except Exception as e:
				print("Mapping error:", e)


if __name__ == "__main__":
	main()


